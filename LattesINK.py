import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import os, threading, csv, openpyxl, time, re

from openpyxl.styles import Font, Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class lattesink:
    def __init__(self):
        self.janela = tk.Tk()
        self.janela.title('Lattes ink')
        #self.janela.iconbitmap('') # Colocar o caminho do arquivo .ico
        self.janela.geometry('800x480')
        self.janela.resizable(width=False, height=False)
        self.janela.configure(bg='#faffff')

        # Estado interno
        self.ids_carregados = []
        self._captura_ativa = False
        self._cancelar_solicitado = False
        self._driver_ativo = None  # Referência ao driver em uso (para fechar ao cancelar)
        self.dados_coletados = []  # Lista de dicionários com os dados de cada pesquisador

        # Frame principal
        self.frame_principal = tk.Frame(self.janela, bg='#faffff')
        self.frame_principal.pack(expand=True, fill='both', padx=20, pady=20)

        # Título
        tk.Label(
            self.frame_principal,
            text='Verificador de Currículos Lattes',
            font=('Arial', 14, 'bold'),
            bg='#faffff'
        ).pack(pady=(0, 15))

        # Seção 1: Pasta de destino
        tk.Label(
            self.frame_principal,
            text='1. Selecione a pasta de destino',
            font=('Arial', 10, 'bold'),
            bg='#faffff',
            anchor='w'
        ).pack(fill='x')

        frame_caminho = tk.Frame(self.frame_principal, bg='white', relief='solid', borderwidth=1)
        frame_caminho.pack(fill='x', pady=(5, 12))

        tk.Label(frame_caminho, text="📁", font=("Arial", 18), bg="white").pack(side="left", padx=10, pady=8)

        self.pasta_destino = tk.StringVar(value="Nenhuma pasta selecionada")
        self.label_caminho = tk.Label(
            frame_caminho,
            textvariable=self.pasta_destino,
            wraplength=1000,
            justify="left",
            bg="white",
            fg="gray",
            font=("Arial", 9, 'bold')
        )
        self.label_caminho.pack(side="left", padx=5, pady=8, fill="x", expand=True)

        tk.Button(
            frame_caminho,
            text="Pasta",
            width=9,
            command=self.escolher_pasta,
            bg="#4CAF50", fg="white",
            font=("Arial", 9, "bold"),
            padx=8, pady=4,
            cursor="hand2", relief='flat'
        ).pack(side="right", padx=10, pady=8)

        # Seção 2: Arquivo de IDs
        tk.Label(
            self.frame_principal,
            text='2. Envie o arquivo com os Lattes IDs (.txt, .xlsx ou .csv)',
            font=('Arial', 10, 'bold'),
            bg='#faffff',
            anchor='w'
        ).pack(fill='x')

        frame_ids = tk.Frame(self.frame_principal, bg='white', relief='solid', borderwidth=1)
        frame_ids.pack(fill='x', pady=(5, 12))

        tk.Label(frame_ids, text="📄", font=("Arial", 18), bg="white").pack(side="left", padx=10, pady=8)

        self.ids_status_var = tk.StringVar(value="Nenhum arquivo selecionado")
        self.label_ids_status = tk.Label(
            frame_ids,
            textvariable=self.ids_status_var,
            justify="left",
            bg="white",
            fg="gray",
            font=("Arial", 9, 'bold')
        )
        self.label_ids_status.pack(side="left", padx=5, pady=8, fill="x", expand=True)

        tk.Button(
            frame_ids,
            text="Lattes ID's",
            width=9,
            command=self.escolher_ids,
            bg="#4CAF50", fg="white",
            font=("Arial", 9, "bold"),
            padx=8, pady=4,
            cursor="hand2", relief='flat'
        ).pack(side="right", padx=10, pady=8)

        # Seção 3: Progresso
        tk.Label(
            self.frame_principal,
            text='3. Progresso da captura',
            font=('Arial', 10, 'bold'),
            bg='#faffff',
            anchor='w'
        ).pack(fill='x')

        frame_progresso = tk.Frame(self.frame_principal, bg='#faffff')
        frame_progresso.pack(fill='x', pady=(5, 12))

        self.progresso_var = tk.DoubleVar(value=0)
        self.barra_progresso = ttk.Progressbar(
            frame_progresso,
            variable=self.progresso_var,
            maximum=100,
            length=560
        )
        self.barra_progresso.pack(side='left', fill='x', expand=True)

        self.label_progresso = tk.Label(
            frame_progresso,
            text="0 / 0",
            font=("Arial", 9, "bold"),
            bg='#faffff',
            width=8
        )
        self.label_progresso.pack(side='left', padx=(8, 0))

        # Log de status
        self.log_var = tk.StringVar(value="Aguardando início...")
        tk.Label(
            self.frame_principal,
            textvariable=self.log_var,
            font=("Arial", 9),
            bg='#faffff',
            fg='#333333',
            anchor='w'
        ).pack(fill='x', pady=(0, 10))

        # Botões de ação (iniciar + cancelar + exportar)
        frame_botoes = tk.Frame(self.frame_principal, bg='#faffff')
        frame_botoes.pack()

        self.btn_iniciar = tk.Button(
            frame_botoes,
            text="Iniciar Captura",
            command=self.iniciar_captura,
            bg="#583f20", fg="white",
            font=("Arial", 10, "bold"),
            padx=10, pady=8,
            cursor="arrow", relief='flat',
            state="disabled"
        )
        self.btn_iniciar.pack(side="left", padx=(0, 8))

        self.btn_cancelar = tk.Button(
            frame_botoes,
            text="Cancelar",
            command=self.cancelar_captura,
            bg="#c0392b", fg="white",
            font=("Arial", 10, "bold"),
            padx=10, pady=8,
            cursor="arrow", relief='flat',
            state="disabled"
        )
        self.btn_cancelar.pack(side="left", padx=(0, 8))
        
        self.btn_exportar = tk.Button(
            frame_botoes,
            text="Exportar Excel",
            command=self.exportar_para_excel,
            bg="#2196F3", fg="white",
            font=("Arial", 10, "bold"),
            padx=10, pady=8,
            cursor="arrow", relief='flat',
            state="disabled"
        )
        self.btn_exportar.pack(side="left")

    # Métodos de UI
    def escolher_pasta(self):
        pasta = filedialog.askdirectory(
            title='Escolha a pasta onde os arquivos serão salvos',
            initialdir=os.path.expanduser('~')
        )
        if pasta:
            self.pasta_destino.set(pasta)
            self.label_caminho.configure(fg="black")
            self._atualizar_btn_iniciar()

    def escolher_ids(self):
        pasta = self.pasta_destino.get()
        if pasta == "Nenhuma pasta selecionada" or not os.path.exists(pasta):
            messagebox.showwarning("Atenção", "Selecione a pasta de destino antes de carregar os IDs.")
            return

        arquivo = filedialog.askopenfilename(
            title='Selecione o arquivo com os Lattes IDs',
            initialdir=os.path.expanduser('~'),
            filetypes=[
                ("Todos os formatos suportados", "*.txt *.xlsx *.csv"),
                ("Arquivo de texto", "*.txt"),
                ("Planilha Excel", "*.xlsx"),
                ("Arquivo CSV", "*.csv"),
                ("Todos os arquivos", "*.*")
            ]
        )
        if not arquivo:
            return

        ext = os.path.splitext(arquivo)[1].lower()

        if ext == '.txt':
            ids = self._ler_ids_txt(arquivo)
            if ids is not None:
                self._aplicar_ids(ids, arquivo)
        elif ext in ('.xlsx', '.csv'):
            self._abrir_seletor_coluna(arquivo, ext)
        else:
            messagebox.showerror("Formato não suportado", "Use arquivos .txt, .xlsx ou .csv.")

    def _ler_ids_txt(self, arquivo):
        """Lê IDs de um arquivo .txt (um por linha). Retorna lista ou None em caso de erro."""
        try:
            with open(arquivo, 'r', encoding='utf-8') as f:
                ids = [linha.strip() for linha in f if linha.strip()]
        except Exception as e:
            messagebox.showerror("Erro ao ler arquivo", f"Não foi possível ler o arquivo:\n{e}")
            return None
        if not ids:
            messagebox.showwarning("Arquivo vazio", "O arquivo não contém nenhum ID.")
            return None
        return ids

    def _aplicar_ids(self, ids, arquivo):
        """Registra a lista de IDs e atualiza a interface."""
        self.ids_carregados = ids
        self.dados_coletados = []  # Reseta os dados coletados
        self.ids_status_var.set(f"✅  {len(ids)} ID(s) carregado(s)  —  {os.path.basename(arquivo)}")
        self.label_ids_status.configure(fg="#2e7d32")
        self.label_progresso.configure(text=f"0 / {len(ids)}")
        self.btn_exportar.configure(state="disabled")
        self._atualizar_btn_iniciar()

    def _ler_colunas_arquivo(self, arquivo, ext):
        """
        Retorna (colunas, dados) onde:
          - colunas: lista de nomes das colunas (str)
          - dados: lista de listas com os valores de cada coluna
        """
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
            linhas = list(ws.iter_rows(values_only=True))
            wb.close()
            if not linhas:
                return [], []
            cabecalho = [str(c) if c is not None else f"Coluna {i+1}"
                         for i, c in enumerate(linhas[0])]
            dados = [[str(row[i]) if row[i] is not None else ""
                      for row in linhas[1:]]
                     for i in range(len(cabecalho))]
            return cabecalho, dados
        else:  # .csv
            with open(arquivo, newline='', encoding='utf-8-sig') as f:
                leitor = csv.reader(f)
                linhas = list(leitor)
            if not linhas:
                return [], []
            cabecalho = [str(c).strip() if c else f"Coluna {i+1}"
                         for i, c in enumerate(linhas[0])]
            dados = [[row[i].strip() if i < len(row) else ""
                      for row in linhas[1:]]
                     for i in range(len(cabecalho))]
            return cabecalho, dados

    def _abrir_seletor_coluna(self, arquivo, ext):
        """Abre uma janela modal para o usuário escolher qual coluna contém os IDs."""
        try:
            colunas, dados = self._ler_colunas_arquivo(arquivo, ext)
        except Exception as e:
            messagebox.showerror("Erro ao ler arquivo", f"Não foi possível ler o arquivo:\n{e}")
            return

        if not colunas:
            messagebox.showwarning("Arquivo vazio", "O arquivo não contém dados.")
            return

        # Janela modal
        modal = tk.Toplevel(self.janela)
        modal.title("Selecionar coluna dos IDs")
        modal.geometry("420x340")
        modal.resizable(False, False)
        modal.configure(bg="#faffff")
        modal.grab_set()  # Bloqueia a janela principal enquanto modal estiver aberta

        tk.Label(
            modal,
            text="Selecione a coluna que contém os Lattes IDs:",
            font=("Arial", 10, "bold"),
            bg="#faffff"
        ).pack(pady=(18, 6), padx=16, anchor="w")

        # Frame com lista + scrollbar
        frame_lista = tk.Frame(modal, bg="#faffff")
        frame_lista.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        scrollbar = tk.Scrollbar(frame_lista)
        scrollbar.pack(side="right", fill="y")

        listbox = tk.Listbox(
            frame_lista,
            yscrollcommand=scrollbar.set,
            font=("Arial", 10),
            selectmode="single",
            activestyle="dotbox",
            height=8
        )
        for col in colunas:
            listbox.insert(tk.END, col)
        listbox.pack(side="left", fill="both", expand=True)
        listbox.selection_set(0)  # pré-seleciona a primeira coluna
        scrollbar.config(command=listbox.yview)

        # Preview dos primeiros valores da coluna selecionada
        lbl_preview = tk.Label(
            modal,
            text="",
            font=("Arial", 8),
            fg="#555",
            bg="#faffff",
            anchor="w",
            wraplength=390
        )
        lbl_preview.pack(padx=16, fill="x")

        def _atualizar_preview(*_):
            sel = listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            amostra = [v for v in dados[idx] if v][:5]
            lbl_preview.configure(
                text="Prévia: " + ",  ".join(amostra) if amostra else "Prévia: (sem dados)"
            )

        listbox.bind("<<ListboxSelect>>", _atualizar_preview)
        _atualizar_preview()  # mostra prévia inicial

        # Botões Confirmar / Cancelar
        frame_btns = tk.Frame(modal, bg="#faffff")
        frame_btns.pack(pady=(4, 14))

        def _confirmar():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Atenção", "Selecione uma coluna.", parent=modal)
                return
            idx = sel[0]
            ids = [v for v in dados[idx] if v]
            if not ids:
                messagebox.showwarning("Coluna vazia",
                    "A coluna selecionada não contém nenhum ID.", parent=modal)
                return
            modal.destroy()
            self._aplicar_ids(ids, arquivo)

        tk.Button(
            frame_btns, text="Confirmar", command=_confirmar,
            bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
            padx=12, pady=6, relief="flat", cursor="hand2"
        ).pack(side="left", padx=(0, 10))

        tk.Button(
            frame_btns, text="Cancelar", command=modal.destroy,
            bg="#c0392b", fg="white", font=("Arial", 10, "bold"),
            padx=12, pady=6, relief="flat", cursor="hand2"
        ).pack(side="left")

    def _atualizar_btn_iniciar(self):
        pasta_ok = (
            self.pasta_destino.get() != "Nenhuma pasta selecionada"
            and os.path.exists(self.pasta_destino.get())
        )
        ids_ok = len(self.ids_carregados) > 0
        if pasta_ok and ids_ok:
            self.btn_iniciar.configure(state="normal", cursor="hand2")
        else:
            self.btn_iniciar.configure(state="disabled", cursor="arrow")

    def _atualizar_log(self, mensagem):
        """Atualiza o log de status na thread principal."""
        self.log_var.set(mensagem)

    def _atualizar_progresso(self, atual, total):
        """Atualiza a barra e o contador na thread principal."""
        self.progresso_var.set((atual / total) * 100)
        self.label_progresso.configure(text=f"{atual} / {total}")

    def _resetar_ids(self):
        """Limpa os IDs carregados e pede que o usuário selecione um novo arquivo."""
        self.ids_carregados = []
        self.dados_coletados = []
        self.ids_status_var.set("Nenhum arquivo selecionado")
        self.label_ids_status.configure(fg="gray")
        self.label_progresso.configure(text="0 / 0")
        self.progresso_var.set(0)
        self.btn_exportar.configure(state="disabled")
        self._atualizar_btn_iniciar()

    def _extrair_dados_pagina(self, driver, lattes_id):
        """
        Extrai nome, ID e data da última atualização da página do Lattes.
        Retorna um dicionário com os dados ou None em caso de erro.
        """
        try:
            # Construir o link do Lattes
            link_lattes = f'https://lattes.cnpq.br/{lattes_id}'

            # Tentar encontrar o nome do pesquisador
            nome = "Não encontrado"
            try:
                # O nome geralmente está em um elemento com classe 'nome' ou em um h1/h2
                elementos_nome = driver.find_elements(By.CSS_SELECTOR, "h1, h2, .nome, .nome-pesquisador")
                for elem in elementos_nome:
                    texto = elem.text.strip()
                    if texto and len(texto) > 3:  # Nome tem pelo menos alguns caracteres
                        nome = texto
                        break
                
                # Se não encontrou, tenta por XPath mais genérico
                if nome == "Não encontrado":
                    xpath_nome = "//div[contains(@class, 'nome')] | //div[contains(@class, 'titulo')] | //h1 | //h2"
                    elementos_nome = driver.find_elements(By.XPATH, xpath_nome)
                    for elem in elementos_nome:
                        texto = elem.text.strip()
                        if texto and len(texto) > 3:
                            nome = texto
                            break
            except Exception:
                pass

            # Tentar encontrar a data da última atualização
            data_atualizacao = "Não encontrada"
            try:
                # Padrões comuns de texto para data de atualização
                padroes = [
                    r"Última atualização[:\s]+(\d{2}/\d{2}/\d{4})",
                    r"atualizado[:\s]+(\d{2}/\d{2}/\d{4})",
                    r"Atualizado[:\s]+(\d{2}/\d{2}/\d{4})",
                    r"Data da última atualização[:\s]+(\d{2}/\d{2}/\d{4})",
                    r"(\d{2}/\d{2}/\d{4})"
                ]
                
                # Pega o texto da página
                page_text = driver.find_element(By.TAG_NAME, "body").text
                
                for padrao in padroes:
                    match = re.search(padrao, page_text)
                    if match:
                        data_atualizacao = match.group(1)
                        break
                        
            except Exception:
                pass

            return {
                'link_lattes': link_lattes,
                'id_lattes': lattes_id,
                'nome': nome,
                'ultima_atualizacao': data_atualizacao,
                'arquivo_screenshot': f'{lattes_id}.png'
            }

        except Exception as e:
            self.janela.after(0, self._atualizar_log, f"⚠️  [{lattes_id}]  Erro ao extrair dados: {e}")
            return {
                'id_lattes': lattes_id,
                'nome': 'Erro na extração',
                'ultima_atualizacao': 'Erro na extração',
                'arquivo_screenshot': f'{lattes_id}.png'
            }

    # Captura de tela
    def iniciar_captura(self):
        """Dispara a captura em uma thread separada para não travar a janela."""
        if self._captura_ativa:
            return

        self._captura_ativa = True
        self._cancelar_solicitado = False
        self.dados_coletados = []  # Reseta os dados coletados
        self.btn_iniciar.configure(state="disabled", cursor="arrow", text="⏳  Capturando...")
        self.btn_cancelar.configure(state="normal", cursor="hand2", text="Cancelar")
        self.btn_exportar.configure(state="disabled")
        self.progresso_var.set(0)

        thread = threading.Thread(target=self._executar_capturas, daemon=True)
        thread.start()

    def cancelar_captura(self):
        """Solicita o cancelamento da captura em andamento."""
        if not self._captura_ativa:
            return

        self._cancelar_solicitado = True
        self.btn_cancelar.configure(state="disabled", cursor="arrow", text="Cancelando...")
        self._atualizar_log("🛑  Cancelamento solicitado — aguardando fechar o navegador atual...")

        # Fecha o driver ativo imediatamente, interrompendo o sleep/wait em curso
        if self._driver_ativo:
            try:
                self._driver_ativo.quit()
            except Exception:
                pass
            self._driver_ativo = None

    def _executar_capturas(self):
        """Roda em background: itera os IDs e chama o Selenium para cada um."""
        total = len(self.ids_carregados)
        pasta = self.pasta_destino.get()
        erros = []

        for i, lattes_id in enumerate(self.ids_carregados, start=1):
            # Verifica cancelamento antes de cada ID
            if self._cancelar_solicitado:
                self.janela.after(0, self._finalizar_cancelamento, i - 1, total)
                return

            self.janela.after(0, self._atualizar_log,
                f"🌐  [{i}/{total}]  Abrindo navegador para ID: {lattes_id}")
            resultado = self._screenshot_e_dados_com_captcha_manual(lattes_id, pasta)

            # Verifica cancelamento logo após fechar o driver
            if self._cancelar_solicitado:
                self.janela.after(0, self._finalizar_cancelamento, i, total)
                return

            if resultado:
                self.dados_coletados.append(resultado)
            else:
                erros.append(lattes_id)
                # Adiciona um registro de erro mesmo assim
                self.dados_coletados.append({
                    'link_lattes': f'https://lattes.cnpq.br/{lattes_id}',
                    'id_lattes': lattes_id,
                    'nome': 'ERRO NA CAPTURA',
                    'ultima_atualizacao': 'ERRO NA CAPTURA',
                    'arquivo_screenshot': f'{lattes_id}.png (erro)'
                })

            self.janela.after(0, self._atualizar_progresso, i, total)

        self.janela.after(0, self._finalizar_captura, erros, total) # Finaliza normalmente na thread principal

    def _screenshot_e_dados_com_captcha_manual(self, lattes_id, pasta_destino):
        """
        Abre o navegador visível, aguarda o usuário resolver o CAPTCHA
        manualmente, extrai os dados da página e salva o screenshot.
        Retorna um dicionário com os dados ou None em caso de erro.
        """
        options = Options()
        options.add_argument("--window-size=1080,720")

        url = f'https://lattes.cnpq.br/{lattes_id}'
        caminho_arquivo = os.path.join(pasta_destino, f'{lattes_id}.png')
        driver = webdriver.Chrome(options=options)
        self._driver_ativo = driver  # Registra para poder fechar ao cancelar

        try:
            driver.get(url)
            self.janela.after(0, self._atualizar_log,
                f"⏳  [{lattes_id}]  Aguardando resolução do CAPTCHA (20s)...")

            # Espera interruptível: verifica o flag a cada 0,5s durante 20s
            for _ in range(40):
                if self._cancelar_solicitado:
                    return None
                time.sleep(0.5)

            # Confirma que a página carregou após o CAPTCHA
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
            except Exception:
                self.janela.after(0, self._atualizar_log,
                    f"⚠️  [{lattes_id}]  Não foi possível confirmar o carregamento da página.")

            if self._cancelar_solicitado:
                return None

            time.sleep(2)  # Aguarda carregamento completo

            # Extrair os dados da página
            dados = self._extrair_dados_pagina(driver, lattes_id)
            
            # Salvar screenshot
            driver.save_screenshot(caminho_arquivo)
            self.janela.after(0, self._atualizar_log,
                f"📸  [{lattes_id}]  Screenshot salvo em: {caminho_arquivo}")
            self.janela.after(0, self._atualizar_log,
                f"📊  [{lattes_id}]  Nome: {dados['nome'][:50]}... | Atualização: {dados['ultima_atualizacao']}")

            return dados

        except Exception as e:
            if not self._cancelar_solicitado:
                self.janela.after(0, self._atualizar_log,
                    f"❌  [{lattes_id}]  Erro: {e}")
            return None

        finally:
            try:
                driver.quit()
            except Exception:
                pass
            self._driver_ativo = None

    def exportar_para_excel(self):
        """Exporta os dados coletados para um arquivo Excel."""
        if not self.dados_coletados:
            messagebox.showwarning("Sem dados", "Nenhum dado foi coletado ainda. Execute a captura primeiro.")
            return

        # Perguntar onde salvar o arquivo Excel
        arquivo_excel = filedialog.asksaveasfilename(
            title='Salvar arquivo Excel',
            defaultextension='.xlsx',
            filetypes=[("Arquivo Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            initialfile='dados_lattes.xlsx'
        )

        if not arquivo_excel:
            return

        try:
            # Criar planilha
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Currículos Lattes"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Cabeçalhos
            cabecalhos = ["Link Lattes", "ID Lattes", "Nome do Pesquisador", "Última Atualização"]
            for col, cabecalho in enumerate(cabecalhos, 1):
                cell = ws.cell(row=1, column=col, value=cabecalho)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Dados
            for row, dados in enumerate(self.dados_coletados, 2):
                ws.cell(row=row, column=1, value=dados['link_lattes'])
                ws.cell(row=row, column=2, value=dados['id_lattes'])
                ws.cell(row=row, column=3, value=dados['nome'])
                ws.cell(row=row, column=4, value=dados['ultima_atualizacao'])
                    
                # Ajustar alinhamento das células de dados
                for col in range(1, 5):
                    ws.cell(row=row, column=col).alignment = center_alignment

            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 45  # Link Lattes
            ws.column_dimensions['B'].width = 20  # ID Lattes
            ws.column_dimensions['C'].width = 40  # Nome
            ws.column_dimensions['D'].width = 20  # Data

            # Salvar arquivo
            wb.save(arquivo_excel)
            
            messagebox.showinfo(
                "Exportação concluída",
                f"Dados exportados com sucesso para:\n{arquivo_excel}\n\nTotal de registros: {len(self.dados_coletados)}"
            )
            
            self._atualizar_log(f"✅  Dados exportados para Excel: {os.path.basename(arquivo_excel)}")

        except Exception as e:
            messagebox.showerror("Erro ao exportar", f"Erro ao salvar arquivo Excel:\n{e}")
            self._atualizar_log(f"❌  Erro ao exportar Excel: {e}")

    def _finalizar_captura(self, erros, total):
        """Chamado na thread principal ao terminar todos os IDs com sucesso."""
        self._captura_ativa = False
        self.btn_iniciar.configure(state="normal", cursor="hand2", text="Iniciar Captura")
        self.btn_cancelar.configure(state="disabled", cursor="arrow", text="Cancelar")
        
        # Habilitar botão de exportar se houver dados
        if self.dados_coletados:
            self.btn_exportar.configure(state="normal", cursor="hand2")

        if erros:
            self._atualizar_log(f"⚠️  Concluído com erros em {len(erros)} ID(s): {', '.join(erros)}")
            messagebox.showwarning(
                "Captura concluída com erros",
                f"{total - len(erros)} de {total} capturas concluídas com sucesso.\n\n"
                f"IDs com erro:\n" + "\n".join(erros)
            )
        else:
            self._atualizar_log(f"✅  Todos os {total} screenshots foram salvos com sucesso!")
            messagebox.showinfo(
                "Captura concluída",
                f"Todos os {total} screenshots foram salvos com sucesso em:\n{self.pasta_destino.get()}\n\n"
                f"Clique em 'Exportar Excel' para gerar a planilha com os dados coletados."
            )

    def _finalizar_cancelamento(self, concluidos, total):
        """Chamado na thread principal quando o usuário cancela."""
        self._captura_ativa = False
        self._cancelar_solicitado = False
        self.btn_cancelar.configure(state="disabled", cursor="arrow", text="Cancelar")
        self.btn_iniciar.configure(state="disabled", cursor="arrow", text="Iniciar Captura")
        
        # Habilitar botão de exportar se houver dados coletados até o cancelamento
        if self.dados_coletados:
            self.btn_exportar.configure(state="normal", cursor="hand2")

        self._atualizar_log(
            f"🛑  Captura cancelada — {concluidos} de {total} ID(s) processados."
        )
        messagebox.showinfo(
            "Captura cancelada",
            f"A captura foi cancelada pelo usuário.\n"
            f"{concluidos} de {total} ID(s) foram processados.\n\n"
            f"Os dados coletados até agora podem ser exportados clicando em 'Exportar Excel'.\n\n"
            f"Selecione um novo arquivo para continuar."
        )
        self._resetar_ids() # Reseta os IDs para forçar nova seleção de arquivo


if __name__ == "__main__":
    app = lattesink()
    app.janela.mainloop()
